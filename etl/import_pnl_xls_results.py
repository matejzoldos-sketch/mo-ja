#!/usr/bin/env python3
"""
Import P&L hodnot z MO-JA report XLSX ("Výsledky" sheet)
→ public.pnl_xls_results_monthly
→ public.pnl_xls_expenses_monthly

Použitie:
  python3 etl/import_pnl_xls_results.py --xlsx-path "../docs/MO-JA_report_2026.xlsx"

Vyžaduje:
  SUPABASE_URL
  SUPABASE_SERVICE_ROLE_KEY (alebo SUPABASE_KEY)
"""

from __future__ import annotations

import argparse
import logging
import os
from decimal import Decimal
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "docs" / "MO-JA_report_2026.xlsx"

log = logging.getLogger("import_pnl_xls_results")

# Section headers / totals — not individual expense line items
SKIP_LABELS = {
    "náklady",
    "obdobie",
    "e-shop",
    "staff",
    "marketing & promo",
    "ostatné",
    "office",
    "podiely zakladateliek",
}


def _to_number(val: object) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float, Decimal)):
        return float(val)
    try:
        s = str(val).strip().replace(" ", "")
        if not s:
            return None
        return float(Decimal(s))
    except Exception:
        return None


def _find_first_row_with_text(ws, text: str, col: int, max_row: int = 2000) -> Optional[int]:
    text_l = text.lower()
    for r in range(1, min(ws.max_row, max_row) + 1):
        v = ws.cell(r, col).value
        if isinstance(v, str) and text_l in v.lower():
            return r
    return None


def _find_costs_row(ws, month_cols: list[int], label_exact: str = "Náklady") -> Optional[int]:
    # "Náklady" sa v sheet-e vyskytuje viac krát; vyberieme ten riadok, kde sú v mesiacoch čísla.
    for r in range(1, min(ws.max_row, 4000) + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.strip().startswith(label_exact):
            has_any_numbers = any(ws.cell(r, c).value is not None for c in month_cols)
            if has_any_numbers:
                return r
    return None


def _find_row_with_text_and_numbers(
    ws,
    needle: str,
    month_cols: list[int],
    max_row: int = 250,
) -> Optional[int]:
    needle_l = needle.lower()
    for r in range(1, min(ws.max_row, max_row) + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and needle_l in v.lower():
            has_any_numbers = any(ws.cell(r, c).value is not None for c in month_cols)
            if has_any_numbers:
                return r
    return None


def _account_prefix(account: object) -> str:
    if account is None:
        return ""
    s = str(int(account)) if isinstance(account, (int, float)) else str(account).strip()
    digits = "".join(ch for ch in s if ch.isdigit())
    return digits[:3] if digits else ""


def extract_expenses(
    ws,
    year_int: int,
    month_cols: list[int],
    month_nums: list[int],
) -> list[dict]:
    """Extract leaf cost rows from Náklady section."""
    start = _find_row_with_text_and_numbers(ws, "e-shop", month_cols) or 38
    end = _find_costs_row(ws, month_cols) or (start + 50)
    if end <= start:
        end = start + 50

    section = "ops"  # ops | staff | marketing
    payload: list[dict] = []

    for r in range(start, end):
        label_raw = ws.cell(r, 2).value
        if not isinstance(label_raw, str) or not label_raw.strip():
            continue
        label = label_raw.strip()
        label_l = label.lower()

        # Section headers
        if label_l == "staff":
            section = "staff"
            continue
        if "marketing" in label_l and "promo" in label_l:
            section = "marketing"
            continue
        if label_l in ("e-shop", "ostatné", "ostatne", "office") or label_l.startswith("podiely"):
            section = "ops"
            continue
        if label_l in SKIP_LABELS:
            continue

        account = ws.cell(r, 1).value
        prefix = _account_prefix(account) or "xls"

        is_staff = section == "staff"
        is_marketing = section == "marketing"

        for month_num, c in zip(month_nums, month_cols):
            amount = _to_number(ws.cell(r, c).value)
            if amount is None or abs(amount) < 0.005:
                continue
            payload.append(
                {
                    "month_key": f"{year_int}-{month_num:02d}",
                    "year": year_int,
                    "supplier": label,
                    "account_prefix": prefix,
                    "amount_eur": round(amount, 2),
                    "is_marketing": bool(is_marketing),
                    "is_staff": bool(is_staff),
                }
            )

    return payload


def extract_results(xlsx_path: Path, sheet_name: str = "Výsledky") -> tuple[list[dict], list[dict]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(
            f"Hárok '{sheet_name}' nie je v {xlsx_path.name}. Dostupné: {', '.join(wb.sheetnames)}"
        )
    ws = wb[sheet_name]

    # Header s mesiacmi býva v riadku 4 a mesiace v stĺpcoch 3..11 (pre template s 9 mesiacmi).
    header_row = 4
    month_cols: list[int] = []
    month_nums: list[int] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, (int, float)) and 1 <= int(v) <= 12:
            month_cols.append(c)
            month_nums.append(int(v))

    if not month_cols:
        # Fallback: month numbers often on row 20/37 in cost/revenue blocks
        for c in range(1, ws.max_column + 1):
            v = ws.cell(37, c).value
            if isinstance(v, (int, float)) and 1 <= int(v) <= 12:
                month_cols.append(c)
                month_nums.append(int(v))

    if not month_cols:
        raise SystemExit("Nepodarilo sa nájsť hlavičku mesiacov v sheet-e (riadok 4).")

    # Year býva v riadku 3 na prvom mesiaci.
    year = _to_number(ws.cell(3, month_cols[0]).value)
    if year is None:
        year = _to_number(ws.cell(36, month_cols[0]).value)
    if year is None:
        raise SystemExit("Nepodarilo sa nájsť rok v sheet-e (riadok 3).")
    year_int = int(year)

    r_revenue = _find_first_row_with_text(
        ws, "Výnosy spolu (v účtovníctve)", col=2
    )
    r_profit_m = _find_first_row_with_text(ws, "Výsledok za mesiac", col=2)
    r_profit_y = _find_first_row_with_text(ws, "Výsledok kumulatívne", col=2)
    r_costs = _find_costs_row(ws, month_cols)
    r_marketing = _find_row_with_text_and_numbers(ws, "Marketing & Promo", month_cols)
    r_opex = _find_row_with_text_and_numbers(ws, "OPEX", month_cols)
    r_other = _find_row_with_text_and_numbers(ws, "Ostatné", month_cols)
    r_staff = _find_row_with_text_and_numbers(ws, "Staff", month_cols)

    if None in (r_revenue, r_costs, r_profit_m, r_profit_y, r_marketing, r_opex, r_other, r_staff):
        raise SystemExit(
            "Nepodarilo sa nájsť riadky pre Výnosy/Náklady/Zisk (skontroluj štruktúru XLS template)."
        )

    monthly: list[dict] = []
    for month_num, c in zip(month_nums, month_cols):
        rev = _to_number(ws.cell(r_revenue, c).value)
        costs = _to_number(ws.cell(r_costs, c).value)
        profit_m = _to_number(ws.cell(r_profit_m, c).value)
        profit_y = _to_number(ws.cell(r_profit_y, c).value)
        marketing = _to_number(ws.cell(r_marketing, c).value)
        opex = _to_number(ws.cell(r_opex, c).value)
        other_operating = _to_number(ws.cell(r_other, c).value)
        staff = _to_number(ws.cell(r_staff, c).value)

        if None in (rev, costs, profit_m, profit_y, marketing, opex, other_operating, staff):
            continue

        month_key = f"{year_int}-{month_num:02d}"
        monthly.append(
            {
                "month_key": month_key,
                "year": year_int,
                "revenue": round(rev, 2),
                "costs": round(costs, 2),
                "profit_month": round(profit_m, 2),
                "profit_ytd": round(profit_y, 2),
                "marketing": round(marketing, 2),
                "opex": round(opex, 2),
                "other_operating": round(other_operating, 2),
                "staff": round(staff, 2),
            }
        )

    expenses = extract_expenses(ws, year_int, month_cols, month_nums)
    wb.close()
    return monthly, expenses


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    ap = argparse.ArgumentParser(description="Import XLS P&L z MO-JA report do Supabase")
    ap.add_argument("--xlsx-path", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--sheet", type=str, default="Výsledky")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    load_dotenv(ROOT / ".env")

    if not args.xlsx_path.is_file():
        raise SystemExit(f"Súbor neexistuje: {args.xlsx_path}")

    sb_url = (os.environ.get("SUPABASE_URL") or "").strip()
    sb_key = (
        os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_KEY") or ""
    ).strip()
    if not sb_url or not sb_key:
        raise SystemExit("Chýba SUPABASE_URL alebo SUPABASE_SERVICE_ROLE_KEY (alebo SUPABASE_KEY) v .env")

    monthly, expenses = extract_results(args.xlsx_path, sheet_name=args.sheet)
    log.info("Načítaných %s mesiacov a %s expense riadkov z XLS.", len(monthly), len(expenses))
    if not monthly:
        return

    if args.dry_run:
        for r in monthly:
            log.info(
                "  %s: revenue=%s costs=%s profit=%s",
                r["month_key"],
                r["revenue"],
                r["costs"],
                r["profit_month"],
            )
        by_supplier: dict[str, float] = {}
        for e in expenses:
            by_supplier[e["supplier"]] = by_supplier.get(e["supplier"], 0) + e["amount_eur"]
        for name, amt in sorted(by_supplier.items(), key=lambda x: -abs(x[1]))[:20]:
            log.info("  expense %-40s %10.2f", name, amt)
        return

    sb = create_client(sb_url, sb_key)
    resp = (
        sb.table("pnl_xls_results_monthly")
        .upsert(monthly, on_conflict="month_key")
        .execute()
    )
    log.info("Monthly upsert OK: %s rows", len(monthly))

    # Replace expenses for imported years
    years = sorted({r["year"] for r in monthly})
    for y in years:
        sb.table("pnl_xls_expenses_monthly").delete().eq("year", y).execute()
    if expenses:
        # upsert in chunks
        chunk = 200
        for i in range(0, len(expenses), chunk):
            part = expenses[i : i + chunk]
            sb.table("pnl_xls_expenses_monthly").upsert(
                part, on_conflict="month_key,supplier,account_prefix"
            ).execute()
    log.info("Expenses upsert OK: %s rows (%s)", len(expenses), resp)


if __name__ == "__main__":
    main()
