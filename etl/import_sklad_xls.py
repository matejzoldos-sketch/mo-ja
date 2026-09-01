#!/usr/bin/env python3
"""
Import fyzického skladu z MO-JA report XLSX (hárok „Sklad_sumár“)
→ public.physical_inventory_monthly

Použitie:
  python3 etl/import_sklad_xls.py --xlsx-path docs/MO-JA_report_2026_31.8.xlsx
  python3 etl/import_sklad_xls.py --dry-run

Vyžaduje: SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY (alebo SUPABASE_KEY)
"""

from __future__ import annotations

import argparse
import logging
import os
from decimal import Decimal
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "docs" / "MO-JA_report_2026.xlsx"
SHEET = "Sklad_sumár"

# Swiss Point „Stav na konci mesiaca“ (R3 hlavička)
PRODUCTS = (
    {
        "product_key": "phase_plus_citron",
        "product_label": "PHASE PLUS - Citrón",
        "in_col": 22,
        "out_col": 23,
        "end_col": 24,
    },
    {
        "product_key": "phase_plus_berry",
        "product_label": "PHASE PLUS - Very Berry",
        "in_col": 25,
        "out_col": 26,
        "end_col": 27,
    },
    {
        "product_key": "phase_ananas",
        "product_label": "PHASE - Ananás",
        "in_col": 28,
        "out_col": 29,
        "end_col": 30,
    },
)

# Výdaj Shopify (R2) — mesačné odpisy z e-shopu
SHOPIFY_OUT_COLS = {
    "phase_plus_citron": 16,
    "phase_plus_berry": 17,
    "phase_ananas": 18,
}

log = logging.getLogger("import_sklad_xls")


def _to_number(val: object) -> Optional[float]:
    if val is None or isinstance(val, bool):
        return None
    if isinstance(val, (int, float, Decimal)):
        return float(val)
    try:
        s = str(val).strip().replace(" ", "")
        if not s:
            return None
        return float(Decimal(s))
    except Exception:
        return None


def _parse_month_num(raw: object) -> Optional[int]:
    if raw is None:
        return None
    if isinstance(raw, (int, float)) and 1 <= int(raw) <= 12:
        return int(raw)
    s = str(raw).strip()
    if s.isdigit() and 1 <= int(s) <= 12:
        return int(s)
    if "-" in s:
        part = s.split("-")[-1].strip()
        if part.isdigit() and 1 <= int(part) <= 12:
            return int(part)
    return None


def extract_sklad(xlsx_path: Path, sheet_name: str = SHEET) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"Hárok {sheet_name!r} v XLS chýba. Dostupné: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows_out: list[dict] = []
    current_year: Optional[int] = None

    for r in range(4, min(ws.max_row, 200) + 1):
        year_val = _to_number(ws.cell(r, 2).value)
        if year_val is None:
            year_val = _to_number(ws.cell(r, 20).value)
        if year_val is not None:
            yi = int(year_val)
            if yi >= 2020:
                current_year = yi
        if current_year is None:
            continue

        month_raw = ws.cell(r, 3).value
        if month_raw is None:
            month_raw = ws.cell(r, 21).value
        month_num = _parse_month_num(month_raw)
        if month_num is None:
            continue

        month_key = f"{current_year}-{month_num:02d}"

        for prod in PRODUCTS:
            stock_end = _to_number(ws.cell(r, prod["end_col"]).value)
            if stock_end is None:
                continue
            stock_in = _to_number(ws.cell(r, prod["in_col"]).value)
            stock_out = _to_number(ws.cell(r, prod["out_col"]).value)
            shopify_col = SHOPIFY_OUT_COLS.get(prod["product_key"])
            shopify_out = (
                _to_number(ws.cell(r, shopify_col).value) if shopify_col else None
            )

            rows_out.append(
                {
                    "month_key": month_key,
                    "year": current_year,
                    "month_num": month_num,
                    "location": "swiss_point",
                    "product_key": prod["product_key"],
                    "product_label": prod["product_label"],
                    "stock_in": stock_in,
                    "stock_out": stock_out,
                    "stock_end": round(stock_end, 2),
                    "shopify_out": shopify_out,
                }
            )

    wb.close()
    return rows_out


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    ap = argparse.ArgumentParser(description="Import Sklad_sumár z MO-JA XLS do Supabase")
    ap.add_argument("--xlsx-path", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--sheet", type=str, default=SHEET)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    load_dotenv(ROOT / ".env")

    if not args.xlsx_path.is_file():
        raise SystemExit(f"Súbor neexistuje: {args.xlsx_path}")

    rows = extract_sklad(args.xlsx_path, sheet_name=args.sheet)
    source_file = str(args.xlsx_path.name)
    for row in rows:
        row["source_file"] = source_file

    log.info("Načítaných %s riadkov (mesiac × produkt) z %s.", len(rows), args.sheet)
    if not rows:
        raise SystemExit("Žiadne dáta — skontroluj štruktúru hárka Sklad_sumár.")

    months = sorted({r["month_key"] for r in rows})
    log.info("Mesiace: %s … %s", months[0], months[-1])

    if args.dry_run:
        latest = months[-1]
        for row in sorted(
            [r for r in rows if r["month_key"] == latest],
            key=lambda x: x["product_key"],
        ):
            log.info(
                "  %s %s: koniec=%s ks (príjem=%s, výdaj=%s, shopify=%s)",
                row["month_key"],
                row["product_label"],
                row["stock_end"],
                row.get("stock_in"),
                row.get("stock_out"),
                row.get("shopify_out"),
            )
        return

    sb_url = (os.environ.get("SUPABASE_URL") or "").strip()
    sb_key = (
        os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_KEY") or ""
    ).strip()
    if not sb_url or not sb_key:
        raise SystemExit("Chýba SUPABASE_URL alebo SUPABASE_SERVICE_ROLE_KEY v .env")

    sb = create_client(sb_url, sb_key)
    chunk = 100
    for i in range(0, len(rows), chunk):
        part = rows[i : i + chunk]
        sb.table("physical_inventory_monthly").upsert(
            part, on_conflict="month_key,location,product_key"
        ).execute()
    log.info("Upsert OK: %s riadkov.", len(rows))


if __name__ == "__main__":
    main()
