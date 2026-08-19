#!/usr/bin/env python3
"""
Import účtovného denníka (CSV) → public.accounting_journal_lines.

Použitie:
  python3 etl/import_accounting_journal_csv.py --csv-path "../docs/Moja - Denník.csv"
  python3 etl/import_accounting_journal_csv.py --csv-path "../docs/MO-JA_report_2026.xlsx" --from-date 2026-06-01

Vyžaduje:
  SUPABASE_URL
  SUPABASE_SERVICE_ROLE_KEY (alebo SUPABASE_KEY)
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import logging
import os
import re
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import List, Optional, Union

from dotenv import load_dotenv
from supabase import create_client


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_CSV = ROOT / "docs" / "Moja - Denník.csv"

load_dotenv(ROOT / ".env")

log = logging.getLogger("import_accounting_journal_csv")

BANK_PAYMENT_RE = re.compile(r"(^|\s)úhrada\s+fp|(^|\s)tb00", re.I)


@dataclass
class JournalRow:
    line_hash: str
    entry_date: str
    month_num: Optional[int]
    doc_number: str
    line_text: str
    debit_account: str
    credit_account: str
    amount_eur: float
    company_name: Optional[str]
    partner_name: Optional[str]
    source_row: int


def parse_amount(raw: Optional[Union[str, int, float]]) -> Optional[float]:
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).replace("€", "").replace("\xa0", " ").strip()
    if not s:
        return None
    s = s.replace(" ", "")
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        left, right = s.rsplit(",", 1)
        if len(right) == 3 and right.isdigit():
            s = left + right
        else:
            s = s.replace(",", ".")
    try:
        return float(Decimal(s))
    except (InvalidOperation, ValueError):
        return None


def parse_date(raw: Optional[Union[str, datetime, date]]) -> Optional[datetime]:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw
    if isinstance(raw, date):
        return datetime(raw.year, raw.month, raw.day)
    s = str(raw).strip()
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def clean_text(val: Optional[str]) -> str:
    return (val or "").strip()


def clean_optional(val: Optional[str]) -> Optional[str]:
    s = clean_text(val)
    return s or None


def line_hash_for(
    entry_date: str,
    doc_number: str,
    debit_account: str,
    credit_account: str,
    amount_eur: float,
    line_text: str,
    partner_name: Optional[str],
    company_name: Optional[str],
    source_row: int,
) -> str:
    # source_row: rovnaké FP riadky (napr. Meta 60 €) musia byť unikátne
    key = "|".join(
        [
            entry_date,
            doc_number,
            debit_account,
            credit_account,
            f"{amount_eur:.4f}",
            line_text,
            partner_name or "",
            company_name or "",
            str(source_row),
        ]
    )
    return hashlib.sha256(key.encode("utf-8")).hexdigest()[:32]


def parse_csv_rows(path: Path) -> List[JournalRow]:
    out: List[JournalRow] = []
    with path.open(encoding="utf-8-sig", errors="replace", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return out

        for row_num, row in enumerate(reader, start=2):
            dt = parse_date(row.get("Dátum"))
            debit = clean_text(row.get("MD"))
            credit = clean_text(row.get("DAL"))
            doc = clean_text(row.get("Číslo"))
            text = clean_text(row.get("Text"))

            if not dt or not debit or not credit or not doc:
                log.warning("Preskočený riadok %s: chýba dátum/účty/doklad", row_num)
                continue

            amount = parse_amount(row.get("Čiastka"))
            if amount is None:
                # Niektoré FP riadky nemajú čiastku na MD (napr. hrubá mzda) — preskočiť.
                continue

            company = clean_optional(row.get("Firma"))
            partner = clean_optional(row.get("Meno"))
            month_raw = clean_text(row.get("Mesiac"))
            month_num = int(month_raw) if month_raw.isdigit() else dt.month
            entry_date = dt.date().isoformat()

            out.append(
                JournalRow(
                    line_hash=line_hash_for(
                        entry_date,
                        doc,
                        debit,
                        credit,
                        amount,
                        text,
                        partner,
                        company,
                        row_num,
                    ),
                    entry_date=entry_date,
                    month_num=month_num,
                    doc_number=doc,
                    line_text=text,
                    debit_account=debit,
                    credit_account=credit,
                    amount_eur=round(amount, 2),
                    company_name=company,
                    partner_name=partner,
                    source_row=row_num,
                )
            )

    return out


def _cell(val: object) -> Optional[Union[str, int, float, datetime, date]]:
    if val is None:
        return None
    if isinstance(val, (datetime, date, int, float)) and not isinstance(val, bool):
        return val
    return str(val)


def parse_xlsx_rows(path: Path, sheet_name: str = "Denník") -> List[JournalRow]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise SystemExit(
            f"Hárok '{sheet_name}' nie je v {path.name}. Dostupné: {', '.join(wb.sheetnames)}"
        )
    ws = wb[sheet_name]
    out: List[JournalRow] = []
    headers: Optional[List[str]] = None

    for row_num, raw in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_num == 1:
            headers = [str(c).strip() if c is not None else "" for c in raw]
            continue
        if not headers:
            continue
        row = {
            headers[i]: _cell(raw[i]) if i < len(raw) else None
            for i in range(len(headers))
            if headers[i]
        }
        dt = parse_date(row.get("Dátum"))
        debit = clean_text(str(row.get("MD") or ""))
        credit = clean_text(str(row.get("DAL") or ""))
        doc = clean_text(str(row.get("Číslo") or ""))
        text = clean_text(str(row.get("Text") or ""))
        if not dt or not debit or not credit or not doc:
            if any(v not in (None, "") for v in (row.get("Dátum"), debit, credit, doc)):
                log.warning("Preskočený riadok %s: chýba dátum/účty/doklad", row_num)
            continue
        amount = parse_amount(row.get("Čiastka"))
        if amount is None:
            continue
        company = clean_optional(str(row.get("Firma") or "") or None)
        partner = clean_optional(str(row.get("Meno") or "") or None)
        month_raw = clean_text(str(row.get("Mesiac") or ""))
        month_num = int(month_raw) if month_raw.isdigit() else dt.month
        entry_date = dt.date().isoformat()
        out.append(
            JournalRow(
                line_hash=line_hash_for(
                    entry_date,
                    doc,
                    debit,
                    credit,
                    amount,
                    text,
                    partner,
                    company,
                    row_num,
                ),
                entry_date=entry_date,
                month_num=month_num,
                doc_number=doc,
                line_text=text,
                debit_account=debit,
                credit_account=credit,
                amount_eur=round(amount, 2),
                company_name=company,
                partner_name=partner,
                source_row=row_num,
            )
        )
    wb.close()
    return out


def load_journal_rows(path: Path, sheet_name: str = "Denník") -> List[JournalRow]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return parse_xlsx_rows(path, sheet_name=sheet_name)
    return parse_csv_rows(path)


def summarize_marketing(rows: List[JournalRow]) -> None:
    from collections import defaultdict

    fees = defaultdict(float)
    excluded = defaultdict(float)
    ads_skip = defaultdict(float)
    unmapped = defaultdict(float)

    def classify(text: str, partner: Optional[str], company: Optional[str], acct: str) -> Optional[str]:
        hay = " ".join([text, partner or "", company or ""]).lower()
        if BANK_PAYMENT_RE.search(hay):
            return None
        if not acct.startswith("518") and not acct.startswith("5015"):
            return None
        if re.search(
            r"shopify|web\s*shop|stripe|visuel|údržba webu|le\s*soft|čechovsk|projektov|ids\s*health|danetax|mof invest|swiss point|lidet|feminea|superfaktura|ytd s|finančný reporting|advokát|matej vida|gs1",
            hay,
        ):
            return "exclude"
        if re.search(r"leri", hay) and re.search(r"konzult", hay):
            return "exclude"
        if re.search(r"meta\s*platforms|meta\s*reklamy", hay):
            return "ads_skip"
        if re.search(
            r"filip|žitňansk|ppc|vk\s*marketing|nastavenie google ads|bartoš|bcreativum|mailer|manychat|canva|agnw|dizajn|kurečkov|ideamaking|copywriting|asaprint|letáky|birne|zelina|promo vide|knižnica|maskér|steli|hup-zagreb|studio echt|upterdam|serica|e-commerce",
            hay,
        ):
            return "fees"
        return "unmapped"

    for r in rows:
        bucket = classify(r.line_text, r.partner_name, r.company_name, r.debit_account)
        if bucket == "fees":
            fees[r.partner_name or r.company_name or "?"] += r.amount_eur
        elif bucket == "exclude":
            excluded[r.partner_name or r.company_name or "?"] += r.amount_eur
        elif bucket == "ads_skip":
            ads_skip[r.partner_name or r.company_name or "?"] += r.amount_eur
        elif bucket == "unmapped":
            unmapped[f"{r.partner_name or r.company_name}: {r.line_text[:40]}"] += r.amount_eur

    log.info("Fees celkom: %.2f EUR (%s dodávateľov)", sum(fees.values()), len(fees))
    for k, v in sorted(fees.items(), key=lambda kv: -kv[1])[:15]:
        log.info("  fees  %10.2f  %s", v, k)
    log.info("Excluded: %.2f EUR", sum(excluded.values()))
    log.info("Ads skip (Meta FP): %.2f EUR", sum(ads_skip.values()))
    if unmapped:
        log.info("Unmapped 518/5015: %.2f EUR", sum(unmapped.values()))
        for k, v in sorted(unmapped.items(), key=lambda kv: -kv[1])[:10]:
            log.info("  ???  %10.2f  %s", v, k)


def upsert_supabase(rows: List[JournalRow], batch_size: int = 500) -> None:
    url = (os.environ.get("SUPABASE_URL") or "").strip()
    key = (
        os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_KEY") or ""
    ).strip()
    if not url or not key:
        raise SystemExit("Chýba SUPABASE_URL alebo SUPABASE_SERVICE_ROLE_KEY v .env")

    sb = create_client(url, key)
    now = datetime.now(timezone.utc).isoformat()

    # Dedup podľa line_hash (PostgREST ON CONFLICT neznesie duplicity v jednom batchi)
    by_hash = {r.line_hash: r for r in rows}
    if len(by_hash) < len(rows):
        log.warning(
            "Deduplikácia: %s → %s unikátnych line_hash",
            len(rows),
            len(by_hash),
        )

    payload = [
        {
            "line_hash": r.line_hash,
            "entry_date": r.entry_date,
            "month_num": r.month_num,
            "doc_number": r.doc_number,
            "line_text": r.line_text,
            "debit_account": r.debit_account,
            "credit_account": r.credit_account,
            "amount_eur": r.amount_eur,
            "company_name": r.company_name,
            "partner_name": r.partner_name,
            "source_row": r.source_row,
            "imported_at": now,
        }
        for r in by_hash.values()
    ]

    for i in range(0, len(payload), batch_size):
        chunk = payload[i : i + batch_size]
        sb.table("accounting_journal_lines").upsert(
            chunk, on_conflict="line_hash"
        ).execute()
        log.info("Upsert %s–%s / %s", i + 1, min(i + batch_size, len(payload)), len(payload))


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

    ap = argparse.ArgumentParser(description="Import účtovného denníka do mo-ja Supabase")
    ap.add_argument("--csv-path", type=Path, default=DEFAULT_CSV)
    ap.add_argument("--sheet", default="Denník", help="Hárok pri xlsx (default Denník)")
    ap.add_argument(
        "--from-date",
        default=None,
        help="Importovať len riadky s dátumom >= YYYY-MM-DD (kvôli posunu source_row)",
    )
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not args.csv_path.is_file():
        raise SystemExit(f"Súbor neexistuje: {args.csv_path}")

    rows = load_journal_rows(args.csv_path, sheet_name=args.sheet)
    if args.from_date:
        rows = [r for r in rows if r.entry_date >= args.from_date]
    dates = [r.entry_date for r in rows]
    log.info("Načítaných %s riadkov z %s", len(rows), args.csv_path)
    if dates:
        log.info("Rozsah dátumov: %s → %s", min(dates), max(dates))

    summarize_marketing(rows)

    if args.dry_run:
        return

    upsert_supabase(rows)


if __name__ == "__main__":
    main()
